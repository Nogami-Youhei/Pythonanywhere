from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse, FileResponse
import numpy as np
import pandas as pd
from sklearn.pipeline import make_pipeline
from sklearn.preprocessing import StandardScaler
from sklearn.linear_model import LinearRegression
from sklearn.decomposition import PCA
from sklearn.cross_decomposition import PLSRegression
import re
import os
import shutil
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
import pandas as pd
from mtranslate import translate
import random
import openpyxl
import openpyxl as xl
from openpyxl.styles import Font
from pathlib import Path
from django.core.files.storage import FileSystemStorage
from .forms import SignupForm, LoginForm
from django.contrib.auth import login, logout
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required


THIS_FOLDER = Path(__file__).parent.resolve()

@login_required
def index(request):
    user = request.user.username
    params = {
        user: user
        }
    return render(request, 'pls/index.html', params)

@login_required
def outline(request):
    return render(request, 'pls/outline.html')

@login_required
def pls(request):
    params = {
        'range1': range(1, 11),
    }
    if (request.method == 'POST'):
        try:
            y = np.array(request.POST['目的変数'].split()).astype('float')
            li = [i.split() for i in request.POST['説明変数'].split('\n')]
            if li[-1] == []:
                li.remove([])
            X = np.array(li).astype('float')
            n_components = int(request.POST['n_components'])

            if X.shape[1] < n_components:
                params = {
                    'range1': range(1, 11),
                    'message': '説明変数の数より主成分の数が多いです！'
                }
                return render(request, 'pls/pls.html', params)

            PLS_model = PLSRegression(n_components=n_components)
            PLS_model.fit(X, y)
            predict = PLS_model.predict(X)
            R2 = round(PLS_model.score(X, y), 4)
            params = {
                'range1': range(1, 11),
                'coef_': 'coef_:<br>' + str(PLS_model.coef_),
                'predict': 'predict:<br>' + str(predict),
                'R2': 'R2: ' + str(R2),
            }
            return render(request, 'pls/pls.html', params)
        except Exception as e:
            params = {
                'range1': range(1, 11),
                'message': e,
            }

    return render(request, 'pls/pls.html', params)

@login_required
def pcr(request):
    params = {
        'range1': range(1, 11),
    }
    if (request.method == 'POST'):
        try:
            y = np.array(request.POST['目的変数'].split()).astype('float')
            li = [i.split() for i in request.POST['説明変数'].split('\n')]
            if li[-1] == []:
                li.remove([])
            X = np.array(li).astype('float')
            n_components = int(request.POST['n_components'])

            if X.shape[1] < n_components:
                params = {
                    'range1': range(1, 11),
                    'message': '説明変数の数より主成分の数が多いです！'
                }
                return render(request, 'pls/pcr.html', params)

            PCR_model = make_pipeline(StandardScaler(),
                            PCA(n_components=n_components),
                            LinearRegression())

            PCR_model.fit(X, y)
            components = PCR_model.named_steps["pca"].components_
            variance_ratio = PCR_model.named_steps["pca"].explained_variance_ratio_
            predict = PCR_model.predict(X)
            R2 = round(PCR_model.score(X, y), 4)
            params = {
                'range1': range(1, 11),
                'components1': 'components:<br>',
                'components2': components,
                'variance_ratio': 'variance_ratio:<br>' + str(variance_ratio),
                'predict': 'predict:<br>' + str(predict),
                'R2': 'R2: ' + str(R2),
            }
            return render(request, 'pls/pcr.html', params)
        except Exception as e:
            params = {
                'range1': range(1, 11),
                'message': e,
            }

    return render(request, 'pls/pcr.html', params)


@login_required
def upload(request):
    if request.method == 'POST' and request.FILES.get('htmlfile'):
        BASE_DIR = Path(__file__).resolve().parent.parent
        media_local_dir = BASE_DIR.joinpath('media_local')
        shutil.rmtree(media_local_dir)
        os.makedirs(media_local_dir, exist_ok=True)
        htmlfile = request.FILES['htmlfile']
        fileobject = FileSystemStorage()
        fileobject.save(htmlfile.name, htmlfile)
        path = media_local_dir.joinpath(htmlfile.name)
        df = pd.read_csv(path, encoding='shift_jis')
        df.iloc[0,0] = 'test'
        df.to_csv(path, header=None, index=None)
        filename, filepath = htmlfile.name, path
        return FileResponse(open(filepath, 'rb'), as_attachment=True, filename=filename)
    
    return render(request, 'pls/upload.html')



def signup_view(request):
    if request.method == 'POST':
        form = SignupForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect(to='/pls/')
        
    else:
        form = SignupForm()
    param = {
        'form': form,
        }
    return render(request, 'pls/signup.html', param)

def login_view(request):
    if request.method == 'POST':
        next = request.POST.get('next')
        form = LoginForm(request, data=request.POST)

        if form.is_valid():
            user = form.get_user()

            if user:
                login(request, user)
                if next == 'None':
                    return redirect(to='/pls/')
                else:
                    return redirect(to=next)

    else:
        form = LoginForm()
        next = request.GET.get('next')

    params = {
        'form': form,
        'next': next,
    }

    return render(request, 'pls/login.html', params)

def logout_view(request):
    logout(request)

    return render(request, 'pls/logout.html')


from .forms import UserForm
from .models import Paper

@login_required
def scraping(request):
    try:
        user = request.user.username
        form = UserForm(request.POST)

        def discriminate_jp_en(string):
            if re.search(r'[ぁ-ん]+|[ァ-ヴー]+|[一-龠]+', string):
                return False
            else:
                return True
        
        if request.method == 'POST':
            if form.is_valid():
                paper = form.save()
                paper.user = user
                paper.save()

                options = webdriver.ChromeOptions()
                user_agent = ['Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/13.0.2 Safari/605.1.15',
                              'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_4) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/13.1 Safari/605.1.15',
                              'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36']

                UA = user_agent[random.randrange(0, len(user_agent), 1)]
                options.add_argument('--user-agent=' + UA)
                options.add_argument('--no-sandbox')
                options.add_argument('--headless')
                options.add_argument("--disable-gpu")

                with webdriver.Chrome(options=options) as driver:
                    ja = paper.ja
                    i = 0
                    result = []
                    keyword = '+'.join(paper.keywords.split())
                    num = paper.number

                    url = 'https://www.jstage.jst.go.jp/result/global/-char/ja?globalSearchKey=' + keyword
                    driver.get(url)
                    html = driver.page_source.encode('utf-8')
                    soup = BeautifulSoup(html, 'html.parser')
                    sortby = paper.choices
                    Select(driver.find_element(By.NAME, 'sortby')).select_by_value(str(sortby))
                    while True:
                        url = driver.current_url
                        html = driver.page_source.encode('utf-8')
                        soup = BeautifulSoup(html, 'html.parser')
                        elems = soup.find('ul', class_='search-resultslisting').find_all('li')
                        if not elems:
                            params = {
                                'message': '検索条件に該当する記事が見つかりません。'
                            }
                            return render(request, 'pls/scraping.html', params)

                        for elem in elems:
                            line = pd.Series(index=['タイトル', 'URL', '学会誌', '出版年', '巻・ページ', '発行日', '公開日', '要約'], dtype=object)
                            line.name = f'{i+1:03d}'
                            anchor = elem.find('a').text
                            if ja and discriminate_jp_en(anchor):
                                anchor = translate(anchor.strip(), 'ja')
                            line[0] = anchor
                            link = elem.find('a')['href']
                            line[1] = link
                            detail = re.split(r'[\n\t]+', elem.find('div', class_='searchlist-additional-info').text)
                            line[2:4] = detail[1:3]
                            line[-3:-1] = detail[-3:-1]
                            line[4] = ', '.join(detail[3:-3])
                            driver.get(link)
                            html = driver.page_source.encode('utf-8')
                            soup = BeautifulSoup(html, 'html.parser')
                            abstract = soup.find('p', class_='global-para-14')
                            if not abstract:
                                line[-1] = ('要約なし')
                            elif abstract.text == '\n':
                                if not abstract.next_sibling:
                                    line[-1] = ('要約なし')
                                elif abstract.next_sibling.text == '\n':
                                    line[-1] = ('要約なし')
                                else:
                                    sibling = abstract.next_sibling
                                    if ja and discriminate_jp_en(sibling.text):
                                        sibling = translate(sibling.text.strip(), 'ja')
                                        line[-1] = sibling.strip()
                                    else:
                                        line[-1] = sibling.text.strip()
                            else:
                                if ja and discriminate_jp_en(abstract.text):
                                    abstract = translate(abstract.text.strip(), 'ja')
                                    line[-1] = abstract.strip()
                                else:
                                    line[-1] = abstract.text.strip()
                            result.append(line)
                            i += 1
                            if i == num:
                                break
                        if i == num:
                            break
                        driver.get(url)
                        ul = driver.find_element(By.XPATH, '//*[@id="search-pagination-wrap-top"]/div/div[1]/ul')
                        button = ul.find_elements(By.TAG_NAME, 'li')[-2]
                        if button.get_attribute('class') == 'inactive-page':
                            break
                        button.click()

                    temp_dir = THIS_FOLDER.joinpath('temp')
                    shutil.rmtree(temp_dir)
                    os.makedirs(temp_dir, exist_ok=True)
                    df = pd.concat(result, axis=1).T
                    file_dir = temp_dir.joinpath(f'{keyword}.xlsx')
                    df.to_excel(file_dir)

                wb = openpyxl.load_workbook(file_dir)
                ws = wb.active
                ws.auto_filter.ref = "A1:I1"
                ws.column_dimensions['B'].width = 20
                ws.column_dimensions['C'].width = 25
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 15
                ws.column_dimensions['F'].width = 15
                ws.column_dimensions['G'].width = 15
                ws.column_dimensions['H'].width = 15
                ws.column_dimensions['I'].width = 50
                wrap_text = xl.styles.Alignment(wrapText=True, vertical='center')

                for row in range(2, ws.max_row+1):
                    ws.row_dimensions[row].height = 100

                for row in ws.iter_rows():
                    for cell in row:
                        cell.alignment = wrap_text

                for cell in ws[f'I2:I{ws.max_row}']:
                    cell[0].alignment = xl.styles.Alignment(wrapText=True, vertical='top')

                for cell in ws['A']:
                    cell.alignment = xl.styles.Alignment(wrapText=True, horizontal='center', vertical='center')

                for cell in ws[f'C2:C{ws.max_row}']:
                    cell[0].hyperlink = cell[0].value
                    cell[0].font = Font(color="0000FF", underline="single")

                wb.save(file_dir)

                filename, filepath = f'{keyword}.xlsx', file_dir
                return FileResponse(open(filepath, 'rb'), as_attachment=True, filename=filename)
        
        else:
            form = UserForm()

    
    except Exception as e:
        params = {
            'message': e,
            }
        return render(request, 'pls/scraping.html', params)

    params = {
        'form': form,
        'name': user,
        }
    return render(request, 'pls/scraping.html', params)


def datalist(request):
    if (request.method == 'POST'):
        data = request.POST.get('data')
        paper = Paper.objects.get(id=int(data))
        print(paper)
        paper.delete()
        return redirect(to='datalist')
    
    papers = Paper.objects.all()
    user = request.user.username

    param = {
        'papers': papers,
        'user': user
    }
    return render(request, 'pls/datalist.html', param)

from .forms import ShapForm
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestRegressor
import shap
import matplotlib.pyplot as plt
from sklearn.model_selection import GridSearchCV
from sklearn.model_selection import KFold
from sklearn.metrics import r2_score
import japanize_matplotlib
from django.utils import timezone 

@login_required
def shap_view(request):
    if request.method == 'POST':
        try:
            clear = request.POST.get('clear')
            if clear:
                path_bar = THIS_FOLDER / 'static' / 'pls' / 'img' / 'bar.png'
                path_dot = THIS_FOLDER / 'static' / 'pls' / 'img' / 'dot.png'
                path_plot = THIS_FOLDER / 'static' / 'pls' / 'img' / 'plot.png'
                os.remove(path_bar)
                os.remove(path_dot)
                os.remove(path_plot)
                return HttpResponse(3)
            
            form = ShapForm(request.POST)
            if form.is_valid():
                target = form.cleaned_data.get('target')
                feature = form.cleaned_data.get('feature')
                row_index = form.cleaned_data.get('row_index')
                n_splits = form.cleaned_data.get('n_splits')

                n_estimators_min = form.cleaned_data.get('n_estimators_min')
                n_estimators_max = form.cleaned_data.get('n_estimators_max')
                n_estimators_div = form.cleaned_data.get('n_estimators_div')
                max_depth_min = form.cleaned_data.get('max_depth_min')
                max_depth_max = form.cleaned_data.get('max_depth_max')
                max_depth_div = form.cleaned_data.get('max_depth_div')
                max_features_min = form.cleaned_data.get('max_features_min')
                max_features_max = form.cleaned_data.get('max_features_max')
                max_features_div = form.cleaned_data.get('max_features_div')

                choices = form.cleaned_data.get('choices')

                def get_params(min_value, max_value, div):
                    if div == 1:
                        return [max_value]
                    result = []
                    interval = (max_value - min_value) / (div - 1)
                    
                    for _ in range(div - 1):
                        result.append(min_value)
                        min_value += interval
                    result.append(max_value)
                    result = list(set(round(i) for i in result))
                    result.sort()
                    
                    return result
                
                n_estimators = get_params(n_estimators_min, n_estimators_max, n_estimators_div)
                max_depth = get_params(max_depth_min, max_depth_max, max_depth_div)
                max_features = get_params(max_features_min, max_features_max, max_features_div)

                li1 = target.split()
                y = pd.DataFrame(li1[1:], columns=[li1[0]], dtype=float)
                li2 = [i.split() for i in feature.split('\n')]
                if li2[-1] == []:
                    li2.remove([])
                X = pd.DataFrame(li2[1:], columns=li2[0], dtype=float)

                if int(choices):
                    model = LinearRegression()
                    scaler = StandardScaler()
                    X_array = scaler.fit_transform(X)
                    model.fit(X_array, y)
                    y_pred = model.predict(X_array)
                    r2 = r2_score(y, y_pred)
                    coef = model.coef_.ravel()
                    X = pd.DataFrame(X_array, columns=li2[0])
                    params = {}
             
                else:
                    rf_grid=RandomForestRegressor()
                    params={'n_estimators': n_estimators,
                            'max_depth'   : max_depth,
                            "max_features": max_features,
                            "random_state": [0]}           

                    y_array = np.array(y).ravel()

                    k_fold=KFold(n_splits=n_splits, shuffle=True, random_state=0)
                    grid=GridSearchCV(estimator=rf_grid, param_grid=params, cv=k_fold, scoring="r2")
                    grid.fit(X, y_array)
                    model = grid.best_estimator_
                    coef = model.feature_importances_
                    params = grid.best_params_
                    r2 = grid.best_score_

                coef_list = [[a, b] for a, b in zip(li2[0], coef)]  

                explainer = shap.Explainer(model, X)
                shap_values = explainer(X)
                shap.summary_plot(shap_values, X, plot_type='bar', show=False)
                path_bar = THIS_FOLDER / 'static' / 'pls' / 'img' / 'bar.png'
                plt.savefig(path_bar)
                plt.close()               

                shap.summary_plot(shap_values, X, plot_type='dot', show=False)
                path_dot = THIS_FOLDER / 'static' / 'pls' / 'img' / 'dot.png'
                plt.savefig(path_dot)
                plt.close()

                shap.waterfall_plot(shap.Explanation(values=shap_values[row_index], data=X.iloc[row_index], feature_names=X.columns), show=False)
                path_plot = THIS_FOLDER / 'static' / 'pls' / 'img' / 'plot.png'
                plt.savefig(path_plot)
                plt.close()
                context = {
                    'form': form,
                    'params': params,
                    'coef_list': coef_list,
                    'r2': r2,
                    'timestamp': timezone.now().timestamp(),
                }
                return render(request, 'pls/shap_result.html', context)
            
        except ValueError as e:
            return HttpResponse(0)

        except IndexError as e:
            return HttpResponse(1)

        except Exception as e:
            return HttpResponse(str(e))

    form = ShapForm()
    context = {
        'form': form
    }
    return render(request, 'pls/shap.html', context)